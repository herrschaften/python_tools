import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import pandas as pd
from PIL import Image
import piexif
import threading
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import tempfile

class ImageExcelGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Image Index Excel Generator")
        self.root.geometry("700x600")
        
        self.image_folder = ""
        self.output_file = ""
        self.image_data = []
        self.temp_files_to_cleanup = []  # Track temp files for cleanup
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        # Input folder frame
        input_frame = ttk.LabelFrame(main_frame, text="Select Image Folder", padding="5")
        input_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        input_frame.columnconfigure(0, weight=1)
        
        self.folder_var = tk.StringVar(value="No folder selected")
        folder_label = ttk.Label(input_frame, textvariable=self.folder_var)
        folder_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        
        browse_folder_btn = ttk.Button(input_frame, text="Browse Folder", command=self.browse_folder)
        browse_folder_btn.grid(row=0, column=1)
        
        # Settings frame
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="5")
        settings_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        settings_frame.columnconfigure(1, weight=1)
        
        ttk.Label(settings_frame, text="Starting Figure Number:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.start_fig_var = tk.IntVar(value=1)
        ttk.Entry(settings_frame, textvariable=self.start_fig_var, width=10).grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(settings_frame, text="Thumbnail Size (pixels):").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        self.thumb_size_var = tk.IntVar(value=100)
        ttk.Entry(settings_frame, textvariable=self.thumb_size_var, width=10).grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
        
        ttk.Label(settings_frame, text="Sort Images By:").grid(row=2, column=0, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        self.sort_method_var = tk.StringVar(value="Creation Date")
        sort_combo = ttk.Combobox(settings_frame, textvariable=self.sort_method_var, values=["Name", "Creation Date"], width=15, state="readonly")
        sort_combo.grid(row=2, column=1, sticky=tk.W, pady=(5, 0))
        
        self.include_subfolders = tk.BooleanVar(value=False)
        ttk.Checkbutton(settings_frame, text="Include subfolders", variable=self.include_subfolders).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # Output info
        output_info = ttk.Label(main_frame, text="Excel file will be created in the same folder as the images", 
                               font=('TkDefaultFont', 8), foreground='gray')
        output_info.grid(row=2, column=0, columnspan=3, pady=(0, 10))
        
        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=3, column=0, columnspan=3, pady=(0, 10))
        
        self.generate_btn = ttk.Button(buttons_frame, text="Generate Excel Index", command=self.start_generation)
        self.generate_btn.grid(row=0, column=0, padx=(0, 5))
        
        preview_btn = ttk.Button(buttons_frame, text="Preview Data", command=self.preview_data)
        preview_btn.grid(row=0, column=1)
        
        # Progress frame
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="5")
        progress_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress_var = tk.StringVar(value="Ready to generate Excel index")
        progress_label = ttk.Label(progress_frame, textvariable=self.progress_var)
        progress_label.grid(row=0, column=0, sticky=tk.W)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Log frame
        log_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding="5")
        log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select folder containing images")
        if folder:
            self.image_folder = folder
            self.folder_var.set(folder)
    
    def log_message(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def get_image_files(self):
        """Get all image files from the selected folder, excluding temporary files"""
        image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
        image_files = []
        
        if self.include_subfolders.get():
            for root, dirs, files in os.walk(self.image_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Check if it's an image file and not a temporary file
                    if (os.path.splitext(file.lower())[1] in image_extensions and 
                        not self.is_temp_file(file_path)):
                        image_files.append(file_path)
        else:
            for file in os.listdir(self.image_folder):
                file_path = os.path.join(self.image_folder, file)
                if (os.path.isfile(file_path) and 
                    os.path.splitext(file.lower())[1] in image_extensions and
                    not self.is_temp_file(file_path)):
                    image_files.append(file_path)
        
        return sorted(image_files)
    
    def is_temp_file(self, file_path):
        """Check if a file is a temporary file that should be excluded"""
        filename = os.path.basename(file_path)
        temp_dir = tempfile.gettempdir()
        
        # Exclude files in temp directory
        if temp_dir in file_path:
            return True
            
        # Exclude files with temp-like names
        temp_indicators = ['tmp', 'temp', 'thumbnail', '_thumb_', '.tmp.', 'temp_']
        filename_lower = filename.lower()
        
        for indicator in temp_indicators:
            if indicator in filename_lower:
                return True
                
        return False
    
    def extract_metadata_from_image(self, filepath):
        """Extract metadata from image (source URL and comment)"""
        metadata = {
            'source_url': '',
            'comment': ''
        }
        
        try:
            if filepath.lower().endswith(('.jpg', '.jpeg')):
                # Extract EXIF data
                exif_data = piexif.load(filepath)
                
                # Extract our custom metadata from UserComment
                if piexif.ExifIFD.UserComment in exif_data.get("Exif", {}):
                    user_comment = exif_data["Exif"][piexif.ExifIFD.UserComment]
                    if isinstance(user_comment, bytes):
                        comment_str = user_comment.decode('utf-8', errors='ignore')
                        # Parse our custom format: "Source: URL | Comment: text"
                        if "Source:" in comment_str:
                            parts = comment_str.split(" | ")
                            for part in parts:
                                if part.startswith("Source:"):
                                    metadata['source_url'] = part.replace("Source:", "").strip()
                                elif part.startswith("Comment:"):
                                    metadata['comment'] = part.replace("Comment:", "").strip()
                
                # Extract from ImageDescription (source URL)
                if piexif.ImageIFD.ImageDescription in exif_data.get("0th", {}):
                    desc = exif_data["0th"][piexif.ImageIFD.ImageDescription]
                    if isinstance(desc, bytes):
                        metadata['source_url'] = desc.decode('utf-8', errors='ignore')
                
                # Extract from Artist field (comment)
                if piexif.ImageIFD.Artist in exif_data.get("0th", {}):
                    artist = exif_data["0th"][piexif.ImageIFD.Artist]
                    if isinstance(artist, bytes):
                        metadata['comment'] = artist.decode('utf-8', errors='ignore')
                        
            elif filepath.lower().endswith('.png'):
                # Extract PNG metadata
                with Image.open(filepath) as img:
                    if hasattr(img, 'info') and img.info:
                        if 'Source URL' in img.info:
                            metadata['source_url'] = img.info['Source URL']
                        if 'Comment' in img.info:
                            metadata['comment'] = img.info['Comment']
                            
        except Exception as e:
            self.log_message(f"Could not read metadata from {os.path.basename(filepath)}: {str(e)}")
        
        return metadata
    
    def create_thumbnail(self, image_path, size):
        """Create a thumbnail of the image for Excel"""
        try:
            with Image.open(image_path) as img:
                # Convert to RGB if necessary (for PNG with transparency)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                # Create thumbnail maintaining aspect ratio
                img.thumbnail((size, size), Image.Resampling.LANCZOS)
                
                # Save to temporary file with a unique name to avoid conflicts
                temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix='thumb_')
                img.save(temp_file.name, 'PNG')
                
                # Track temp file for cleanup
                self.temp_files_to_cleanup.append(temp_file.name)
                
                return temp_file.name
        except Exception as e:
            self.log_message(f"Could not create thumbnail for {os.path.basename(image_path)}: {str(e)}")
            return None
    
    def process_images(self):
        """Process all images and extract data"""
        image_files = self.get_image_files()
        
        if not image_files:
            return []
        
        self.log_message(f"Found {len(image_files)} valid image files (excluding temp files)")
        
        self.progress_bar['maximum'] = len(image_files)
        self.progress_bar['value'] = 0
        
        data = []
        start_fig = self.start_fig_var.get()
        
        # Sort images based on user selection
        sort_method = self.sort_method_var.get()
        
        if sort_method == "Creation Date":
            # Create list of image files with their creation dates
            image_data_with_dates = []
            
            for image_path in image_files:
                try:
                    # Get creation date
                    stat = os.stat(image_path)
                    creation_date = stat.st_ctime  # Creation time (timestamp)
                    image_data_with_dates.append((image_path, creation_date))
                except Exception as e:
                    self.log_message(f"Could not get creation date for {os.path.basename(image_path)}: {str(e)}")
                    # Add with current time if can't get creation date
                    image_data_with_dates.append((image_path, 0))
            
            # Sort by creation date (oldest first)
            image_data_with_dates.sort(key=lambda x: x[1])
            sorted_image_files = [item[0] for item in image_data_with_dates]
            self.log_message(f"Sorted {len(sorted_image_files)} images by creation date (oldest first)")
            
        else:  # Sort by name
            sorted_image_files = sorted(image_files, key=lambda x: os.path.basename(x).lower())
            self.log_message(f"Sorted {len(sorted_image_files)} images by name (alphabetical)")
        
        # Process sorted images
        for i, image_path in enumerate(sorted_image_files):
            self.progress_var.set(f"Processing {i+1}/{len(sorted_image_files)}: {os.path.basename(image_path)}")
            self.progress_bar['value'] = i + 1
            self.root.update_idletasks()
            
            # Extract metadata
            metadata = self.extract_metadata_from_image(image_path)
            
            # Get filename without extension for caption
            filename = os.path.basename(image_path)
            name_without_ext = os.path.splitext(filename)[0]
            
            # Create thumbnail
            thumbnail_path = self.create_thumbnail(image_path, self.thumb_size_var.get())
            
            # Prepare row data
            row_data = {
                'chapter': '',  # Empty as requested
                'thumbnail_path': thumbnail_path,  # We'll handle this separately
                'name': filename,
                'prefix': 'Figure',
                'fig_number': start_fig + i,
                'sub_number': '',  # Will be filled by formula
                'caption': name_without_ext,
                'type': metadata['comment'] if metadata['comment'] else '',
                'link': metadata['source_url'] if metadata['source_url'] else '',
                'access': '',  # Will be filled by formula
                'full_citation': ''  # Will be filled by formula
            }
            
            data.append(row_data)
            self.log_message(f"✓ Processed: {filename}")
        
        return data
    
    def create_excel_file(self, data):
        """Create the Excel file with formulas and thumbnails"""
        if not data:
            self.log_message("ERROR: No data provided to create_excel_file")
            return None
        
        self.log_message(f"Creating Excel file for {len(data)} images...")
        
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"ImageIndex_{timestamp}.xlsx"
            output_path = os.path.join(self.image_folder, output_filename)
            self.log_message(f"Output path: {output_path}")
            
            # Create workbook and worksheet
            self.log_message("Creating workbook...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ImageIndex"
            
            # Define headers
            headers = ['chapter', 'thumbnail', 'name', 'prefix', 'fig_number', 'sub_number', 
                      'caption', 'type', 'link', 'access', 'full_citation']
            
            # Write headers
            self.log_message("Writing headers...")
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data and add formulas
            self.log_message("Writing data and formulas...")
            for row_idx, row_data in enumerate(data, 2):
                try:
                    # Basic data
                    ws.cell(row=row_idx, column=1, value=row_data['chapter'])  # chapter (empty)
                    # Skip thumbnail for now (column 2)
                    ws.cell(row=row_idx, column=3, value=row_data['name'])  # name
                    ws.cell(row=row_idx, column=4, value=row_data['prefix'])  # prefix
                    ws.cell(row=row_idx, column=5, value=row_data['fig_number'])  # fig_number
                    
                    # Updated sub_number formula (column 6) - using your specified formula
                    sub_formula = f'IF(E{row_idx}=E{row_idx+1},CHAR(97+COUNTIF($E$1:E{row_idx},E{row_idx})-COUNTIF($E$1:$E$1,E{row_idx})-1),IF(E{row_idx}=E{row_idx-1},CHAR(97+COUNTIF($E$1:E{row_idx},E{row_idx})-COUNTIF($E$1:$E$1,E{row_idx})-1),""))'
                    ws.cell(row=row_idx, column=6, value=f'={sub_formula}')
                    
                    ws.cell(row=row_idx, column=7, value=row_data['caption'])  # caption
                    ws.cell(row=row_idx, column=8, value=row_data['type'])  # type
                    ws.cell(row=row_idx, column=9, value=row_data['link'])  # link
                    
                    # Access date - static current date when Excel is generated
                    if row_data['link']:  # Only add access date if there's a link
                        current_date = datetime.now().strftime("accessed %B %d, %Y")
                        ws.cell(row=row_idx, column=10, value=current_date)
                    else:
                        ws.cell(row=row_idx, column=10, value="")  # Empty if no link
                    
                    # Full citation formula (column 11) - simplified
                    citation_formula = f'D{row_idx} & " " & E{row_idx} & IF(F{row_idx}<>"",F{row_idx},"") & ". " & G{row_idx} & ". " & H{row_idx} & IF(I{row_idx}<>"", " " & I{row_idx} & IF(J{row_idx}<>"", ", " & J{row_idx},""),"") & "."'
                    ws.cell(row=row_idx, column=11, value=f'={citation_formula}')
                    
                except Exception as row_error:
                    self.log_message(f"ERROR writing row {row_idx}: {str(row_error)}")
                    continue
            
            # Set column widths
            self.log_message("Setting column widths...")
            column_widths = [10, 15, 25, 10, 12, 12, 30, 20, 40, 25, 60]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Set row heights for thumbnails
            self.log_message("Setting row heights...")
            for row_idx in range(2, len(data) + 2):
                ws.row_dimensions[row_idx].height = self.thumb_size_var.get() * 0.75  # Adjust for Excel units
            
            # Save workbook first (without thumbnails)
            self.log_message("Saving initial workbook...")
            wb.save(output_path)
            
            # Add thumbnails in a separate step
            self.log_message("Adding thumbnails...")
            thumbnail_count = 0
            
            for row_idx, row_data in enumerate(data, 2):
                if row_data['thumbnail_path'] and os.path.exists(row_data['thumbnail_path']):
                    try:
                        # Verify the thumbnail file exists and is valid
                        thumbnail_size = os.path.getsize(row_data['thumbnail_path'])
                        self.log_message(f"Processing thumbnail: {row_data['thumbnail_path']} (size: {thumbnail_size} bytes)")
                        
                        img = XLImage(row_data['thumbnail_path'])
                        
                        # Resize image to fit cell exactly
                        cell_width_pixels = self.thumb_size_var.get()
                        cell_height_pixels = self.thumb_size_var.get()
                        
                        img.width = cell_width_pixels
                        img.height = cell_height_pixels
                        
                        # Use TwoCellAnchor with "twoCell" editAs for "Move and size with cells" behavior
                        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
                        
                        # Calculate cell coordinates (0-indexed)
                        col_idx = 1  # Column B (0-indexed)
                        row_idx_0 = row_idx - 1  # 0-indexed row
                        
                        # Create anchor markers with small offsets to keep image within cell
                        offset = 5000  # Small offset in EMUs to keep image inside cell
                        
                        _from = AnchorMarker(
                            col=col_idx, 
                            row=row_idx_0, 
                            colOff=offset, 
                            rowOff=offset
                        )
                        
                        to = AnchorMarker(
                            col=col_idx, 
                            row=row_idx_0, 
                            colOff=cell_width_pixels * 9525 - offset,  # Convert pixels to EMUs
                            rowOff=cell_height_pixels * 9525 - offset
                        )
                        
                        # Use TwoCellAnchor with editAs="twoCell" for proper cell binding
                        img.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
                        
                        ws.add_image(img)
                        thumbnail_count += 1
                        
                        self.log_message(f"✓ Added thumbnail {thumbnail_count} at row {row_idx} with TwoCellAnchor")
                        
                    except Exception as thumb_error:
                        self.log_message(f"ERROR adding thumbnail for row {row_idx}: {str(thumb_error)}")
                        # Add placeholder text if image fails
                        ws.cell(row=row_idx, column=2, value="[Thumbnail Error]")
                        
                        # Try to diagnose the issue
                        if os.path.exists(row_data['thumbnail_path']):
                            file_size = os.path.getsize(row_data['thumbnail_path'])
                            self.log_message(f"  Thumbnail file exists, size: {file_size} bytes")
                        else:
                            self.log_message(f"  Thumbnail file does not exist: {row_data['thumbnail_path']}")
            
            # Save final workbook with thumbnails
            self.log_message(f"Saving final workbook with {thumbnail_count} thumbnails...")
            wb.save(output_path)
            
            # Clean up temporary files AFTER saving
            self.cleanup_temp_files()
            
            self.log_message(f"SUCCESS: Excel file created at {output_path}")
            return output_path
            
        except Exception as e:
            self.log_message(f"ERROR in create_excel_file: {str(e)}")
            import traceback
            self.log_message(f"Full traceback: {traceback.format_exc()}")
            return None
    
    def cleanup_temp_files(self):
        """Clean up all temporary thumbnail files"""
        self.log_message("Cleaning up temporary thumbnail files...")
        cleanup_count = 0
        
        for temp_file in self.temp_files_to_cleanup:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    cleanup_count += 1
            except Exception as cleanup_error:
                self.log_message(f"Could not cleanup temp file {temp_file}: {str(cleanup_error)}")
        
        self.log_message(f"Cleaned up {cleanup_count} temporary files")
        self.temp_files_to_cleanup.clear()
    
    def preview_data(self):
        """Preview the data that would be generated"""
        if not self.image_folder:
            messagebox.showwarning("No Folder", "Please select an image folder first.")
            return
        
        # Clear log
        self.log_text.delete('1.0', tk.END)
        
        self.log_message("Previewing data...")
        image_files = self.get_image_files()
        
        if not image_files:
            self.log_message("No image files found in the selected folder.")
            return
        
        self.log_message(f"Found {len(image_files)} image files:")
        self.log_message("-" * 50)
        
        start_fig = self.start_fig_var.get()
        
        for i, image_path in enumerate(image_files[:10]):  # Preview first 10
            filename = os.path.basename(image_path)
            name_without_ext = os.path.splitext(filename)[0]
            metadata = self.extract_metadata_from_image(image_path)
            
            self.log_message(f"Figure {start_fig + i}: {filename}")
            self.log_message(f"  Caption: {name_without_ext}")
            self.log_message(f"  Type: {metadata['comment'] if metadata['comment'] else '(no comment)'}")
            self.log_message(f"  Link: {metadata['source_url'] if metadata['source_url'] else '(no source URL)'}")
            self.log_message("")
        
        if len(image_files) > 10:
            self.log_message(f"... and {len(image_files) - 10} more files")
    
    def generation_worker(self):
        """Worker function for generating Excel file"""
        try:
            self.log_message("Starting image processing...")
            self.progress_var.set("Processing images...")
            data = self.process_images()
            
            self.log_message(f"Processed {len(data)} images")
            
            if not data:
                self.log_message("ERROR: No data to process")
                self.root.after(0, lambda: messagebox.showwarning("No Data", "No image files found or processed."))
                return
            
            self.log_message("Creating Excel file...")
            self.progress_var.set("Creating Excel file...")
            
            try:
                output_path = self.create_excel_file(data)
                self.log_message(f"Excel file creation returned: {output_path}")
                
                if output_path and os.path.exists(output_path):
                    self.log_message(f"SUCCESS: Excel file created at {output_path}")
                    self.root.after(0, lambda: self.generation_complete(output_path, len(data)))
                else:
                    self.log_message("ERROR: Excel file creation failed - no output path or file doesn't exist")
                    self.root.after(0, lambda: messagebox.showerror("Error", "Failed to create Excel file - check log for details"))
                    
            except Exception as excel_error:
                self.log_message(f"ERROR in Excel creation: {str(excel_error)}")
                import traceback
                self.log_message(f"Full traceback: {traceback.format_exc()}")
                self.root.after(0, lambda: messagebox.showerror("Excel Error", f"Excel creation failed: {str(excel_error)}"))
                
        except Exception as e:
            self.log_message(f"ERROR in generation worker: {str(e)}")
            import traceback
            self.log_message(f"Full traceback: {traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {str(e)}"))
        finally:
            self.root.after(0, self.reset_ui)
    
    def generation_complete(self, output_path, count):
        """Called when generation is complete"""
        self.progress_var.set(f"Complete! Generated index for {count} images")
        self.log_message(f"\n=== Generation Complete ===")
        self.log_message(f"Excel file created: {output_path}")
        self.log_message(f"Total images processed: {count}")
        
        messagebox.showinfo("Success", 
                           f"Excel index created successfully!\n\n"
                           f"File: {os.path.basename(output_path)}\n"
                           f"Location: {os.path.dirname(output_path)}\n"
                           f"Images processed: {count}")
    
    def reset_ui(self):
        """Reset UI after generation"""
        self.generate_btn.config(state="normal")
        self.progress_bar['value'] = 0
    
    def start_generation(self):
        """Start the Excel generation process"""
        if not self.image_folder:
            messagebox.showwarning("No Folder", "Please select an image folder first.")
            return
        
        # Clean up any existing temp files before starting
        self.cleanup_temp_files()
        
        # Clear log
        self.log_text.delete('1.0', tk.END)
        
        # Disable button and start process
        self.generate_btn.config(state="disabled")
        self.progress_var.set("Starting generation...")
        
        # Start worker thread
        thread = threading.Thread(target=self.generation_worker)
        thread.daemon = True
        thread.start()

def main():
    root = tk.Tk()
    app = ImageExcelGenerator(root)
    root.mainloop()

if __name__ == "__main__":
    main()